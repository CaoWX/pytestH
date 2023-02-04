# coding:utf-8
# Author:窝里横
import os

import yaml
import json
from configparser import ConfigParser
from common.logger import logger
import xlrd
import openpyxl



class MyConfigParser(ConfigParser):
    # 重写 configparser 中的 optionxform 函数，解决 .ini 文件中的 键option 自动转为小写的问题
    def __init__(self, defaults=None):
        ConfigParser.__init__(self, defaults=defaults)

    def optionxform(self, optionstr):
        return optionstr

class ReadFileData():

    def __init__(self):
        pass

    def load_yaml(self, file_path):
        logger.info("加载 {} 文件......".format(file_path))
        with open(file_path, encoding='utf-8') as f:
            data = yaml.safe_load(f)
        logger.info("读到数据 ==>>  {} ".format(data))
        return data

    def load_json(self, file_path):
        logger.info("加载 {} 文件......".format(file_path))
        with open(file_path, encoding='utf-8') as f:
            data = json.load(f)
        logger.info("读到数据 ==>>  {} ".format(data))
        return data

    def load_ini(self, file_path):
        logger.info("加载 {} 文件......".format(file_path))
        config = MyConfigParser()
        config.read(file_path, encoding="UTF-8")
        data = dict(config._sections)
        # print("读到数据 ==>>  {} ".format(data))
        return data

class DoExcel:
    def __init__(self, filepath):
        self.path = filepath
        print('正在使用xlrd读取%s的数据' % self.path)

    # 使用xlrd读取数据
    def get_xlrddata(self):
        all_data = []
        file_data = xlrd.open_workbook(self.path)
        table = file_data.sheets()[0]
        nrows = table.nrows
        for row in range(1, nrows):
            all_data.append(
                {
                    '用例ID': table.cell(row, 0).value,  # 表格index从0开始
                    '用例名称': table.cell(row, 1).value,
                    '用例优先级': table.cell(row, 2).value,
                    'url': table.cell(row, 3).value,
                    'params': eval(table.cell(row, 4).value),
                    '请求方式': table.cell(row, 5).value,
                    '返回': table.cell(row, 6).value,
                    '是否通过': table.cell(row, 7).value
                }
            )
        return all_data
    #使用xlwt写如数据,。。。有点问题
    # def write_xlrddata(self):
    #     old_excel = xlrd.open_workbook(self.path)
    #     new_excel = copy(old_excel)
    #     ws = new_excel.sheets()[0]
    #     try:
    #         ws.cell(1, 7).value = 'name'
    #
    #     except Exception as e:
    #         print('写入文件出错，出错原因%s' % e)



    #使用openxlpy 读写数据
    def openpyxl_data(self):
        all_data = []
        file = openpyxl.load_workbook(self.path)
        table = file['Sheet1']

        for i in range(1, len(list(table.rows))):
            all_data.append(
                {
                    'ID': list(table.rows)[i][0].value,  # 表格index从1开始
                    'Name': list(table.rows)[i][1].value,
                    '用例优先级': list(table.rows)[i][2].value,
                    'url': list(table.rows)[i][3].value,
                    'params': eval(list(table.rows)[i][4].value),
                    'RequestMethod': list(table.rows)[i][5].value,
                    'ExpectedResult': list(table.rows)[i][6].value,
                    '是否通过': list(table.rows)[i][5].value

                }
            )
        return all_data

    def write_openxlpy(self,row, column):
        workbool = openpyxl.load_workbook(self.path)
        print(workbool)
        table = workbool['Sheet1']
        print(table)
        try:
            table.cell(row=row,column=column, value='pass')

        except Exception as  e:
            print('写入文件出错，出错原因%s' % e)

        finally:
            workbool.save(self.path)

data = ReadFileData()
BASE_PATH = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
sql_data_file_path = os.path.join(BASE_PATH, "config", "api_demoProject.conf")
sqldata = data.load_ini(sql_data_file_path)["MYSQL"]

api_data_file_path = os.path.join(BASE_PATH, "data", "api_test_data.yml")
apidata  = data.load_yaml(api_data_file_path)
print(apidata['test_login_data'])
